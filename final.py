import tkinter
from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk

def recall(): #
    def close():
        count_item.quit()
        count_item.destroy()
    def go():
        global temp_sheet2
        global room

        print(amount.get())
        print(type(amount.get()))

        def setroom():
            빈소 = tkinter.Label(win, text=room, width=35, height=2, relief="solid")
            빈소.place(x=50, y=10)
        def setpinfo(num):

            상주성명 = tkinter.Label(win, text=pinfo_sheet.cell(num, 1).value, width=35, height=2, relief="solid")
            상주성명.place(x=50, y=60)

            빈소기간 = tkinter.Label(win, text=pinfo_sheet.cell(num, 2).value, width=35, height=2, relief="solid")
            빈소기간.place(x=50, y=110)

        if amount.get()=="0":
            temp_sheet2=info_sheets[0]
            room="빈소1"
            setroom()
            setpinfo(1)
        elif amount.get()=="1":
            temp_sheet2=info_sheets[1]
            room = "빈소2"
            setroom()
            setpinfo(2)
        elif amount.get()=="2":
            temp_sheet2=info_sheets[2]
            room = "빈소3"
            setroom()
            setpinfo(3)
        elif amount.get()=="3":
            temp_sheet2=info_sheets[3]
            room = "빈소5"
            setroom()
            setpinfo(4)
            print(temp_sheet2)
        elif amount.get()=="4":
            temp_sheet2=info_sheets[4]
            room = "빈소6"
            setroom()
            setpinfo(5)
        elif amount.get()=="5":
            temp_sheet2=info_sheets[5]
            room = "특101"
            setroom()
            setpinfo(6)
        elif amount.get()=="6":
            temp_sheet2=info_sheets[6]
            room = "특102"
            setroom()
            setpinfo(7)
        elif amount.get()=="7":
            temp_sheet2=info_sheets[7]
            room = "특201"
            setroom()
            setpinfo(8)
        elif amount.get()=="8":
            temp_sheet2=info_sheets[8]
            room = "특202"
            setroom()
            setpinfo(9)

        print(amount.get())
        insert_tree2(temp_sheet2)
        close()



    count_item = Tk()  # 불러오기 하면 나오는 화면

    count_item.geometry("200x150+500+300")  # 창의 크기
    count_item.title("수량 입력")  # 창의 제목
    count_item.option_add("*Font", "맑은고딕 14")  # 전체 폰트

    ontk = Label(count_item)  # 수량 레이블
    ontk.config(text="수량", width=10, relief="solid")
    ontk.pack(side="top", pady=10)

    amount = Entry(count_item)  # 수량 엔트리 go_enter 연결
    amount.config(width=10, relief="solid", borderwidth=0)
    amount.focus()
    # amount.bind("<Return>", go_enter)
    amount.place(x=60, y=50)
    amount.pack()

    conf = Button(count_item, text="확인")  # 확인 버튼
    conf.config(width=10, height=3, command=go)  # go 연결
    # conf.place(x=30,y=200)
    conf.pack(side="bottom", pady=10)
    count_item.mainloop()

def del_t(): #오른쪽 트리 삭제용
    tree.delete(*tree.get_children())
def del_t2(): #오른쪽 트리 삭제용
    tree2.delete(*tree2.get_children())

def btn1():
    insert_tree(og_sheets[0])
    global temp_sheet
    temp_sheet = og_sheets[0]
def btn2():
    insert_tree(og_sheets[1])
    global temp_sheet
    temp_sheet = og_sheets[1]
def btn3():
    insert_tree(og_sheets[2])
    global temp_sheet
    temp_sheet = og_sheets[2]
def btn4():
    insert_tree(og_sheets[3])
    global temp_sheet
    temp_sheet = og_sheets[3]
def btn5():
    insert_tree(og_sheets[4])
    global temp_sheet
    temp_sheet=og_sheets[4]

def insert_tree(sheet): #자료형 변환해서 화면 표시, 저장
    del_t()
    print("inset_tree()")
    row = []
    modified_sheet = []

    for x in range(2, (sheet.max_row + 1)):
        for y in range(1, 5):
            if (sheet.cell(x, 1).value == None) | (sheet.cell(x, 1).value == '') | (
                    sheet.cell(x, 1).value == 0):  # 물품명이 None, '', 0 이면 참조 끝
                break
            elif (y == 4) & (sheet.cell(x, 4).value == None):  # None이면 빈셀로
                row.append(" ")
            elif sheet.cell(x, y).value == None:  # None이면 0으로
                row.append(0)
            elif (y != 1) & (y != 4) & (type(sheet.cell(x, y).value) == str):  # 물품명이 아니면서 str일 경우
                print(1)
                row.append(int(float(sheet.cell(x, y).value)))
            else:
                row.append(sheet.cell(x, y).value)
        modified_sheet.append(row)
        row = []
        tree.insert('', 'end', text="", values=modified_sheet[x - 2])
    og_file.save(home)
def l_click(event):
    def close():
        count_item.quit()
        count_item.destroy()
    def go(): #확인 버튼
        num=int(amount.get()) #입력된 텍스트(수량)저장
        selectedItem = tree.selection()[0]  # tree 선택한 위치 받기
        onoff=True

        row=[] #지역변수 리셋 필요 없음
        if (((tree.item(selectedItem)['values'][2])==None) | (tree.item(selectedItem)['values'][2]<num)):
            messagebox.showinfo("","수량보다 많이 입력하였습니다.")
        else:
            # for row2 in temp_sheet.iter_rows(min_row=2):
            #     print(row2[0].value)
            for row2 in temp_sheet.iter_rows(min_row=2): #왼쪽 수량 조절
                for cell in row2:
                    print("cell in row2")
                    if cell.value == tree.item(selectedItem)['values'][0]:
                        row2[2].value = int(tree.item(selectedItem)['values'][2])-num
                        og_file.save(home)
                        print("og_file.save(home)")
                        break
            for row3 in temp_sheet2.iter_rows(): #중앙 수량 조절: 이름이 겹치면 실행. 추가 안하고 특정 값만 수정
                for cell in row3:
                    if cell.value==tree.item(selectedItem)['values'][0]:
                        row3[2].value +=num
                        row3[3].value=row3[1].value*row3[2].value
                        onoff=False #물품명이 겹치면 아래의 (if onoff=True:)를 사용 안함
                        break
            if onoff==True: #중앙 수량 조절: 이름이 안겹칠때 실행. 새로 물품을 추가
                temp_sheet2.append([tree.item(selectedItem)['values'][0],tree.item(selectedItem)['values'][1],num,(tree.item(selectedItem)['values'][1]*num),tree.item(selectedItem)['values'][3]])
            # for rows in info_sheets[0].iter_rows(min_row=1):
            #     info_sheets[0].delete_rows(0)
            info_file.save(info_xl) #오른쪽 시트 저장
        insert_tree2(temp_sheet2) #오른쪽에 물건 표시
        l_refrech() #왼쪽 수량 변화용 리프레시
        close()
    def go_enter(event): #확인 버튼
        num=int(amount.get()) #입력된 텍스트(수량)저장
        selectedItem = tree.selection()[0]  # tree 선택한 위치 받기
        onoff=True
        row=[] #지역변수 리셋 필요 없음
        if(((tree.item(selectedItem)['values'][2])==None)| (tree.item(selectedItem)['values'][2]<num)):
            messagebox.showinfo("","수량보다 많이 입력하였습니다.")
        else:
            for row2 in temp_sheet.iter_rows(min_row=2): #왼쪽 수량 조절
                for cell in row2:
                    print("cell in row2")
                    if cell.value == tree.item(selectedItem)['values'][0]:
                        row2[2].value = int(tree.item(selectedItem)['values'][2])-num
                        og_file.save(home)
                        print("og_file.save(home)")
                        break

            for row3 in temp_sheet2.iter_rows(): #중앙 수량 조절
                for cell in row3:
                    if cell.value==tree.item(selectedItem)['values'][0]:
                        row3[2].value +=num
                        row3[3].value=row3[1].value*row3[2].value
                        onoff=False
                        break

            if onoff==True:
                temp_sheet2.append([tree.item(selectedItem)['values'][0],tree.item(selectedItem)['values'][1],num,(tree.item(selectedItem)['values'][1]*num),tree.item(selectedItem)['values'][3]])

            info_file.save(info_xl) #오른쪽 시트 저장
        insert_tree2(temp_sheet2) #오른쪽에 물건 표시
        l_refrech() #왼쪽 수량 변화용 리프레시
        close()

    global temp_sheet
    print(temp_sheet)
    print(temp_sheet2)
    count_item = Tk()  # 불러오기 하면 나오는 화면

    count_item.geometry("200x150+500+300")  # 창의 크기
    count_item.title("수량 입력")  # 창의 제목
    count_item.option_add("*Font", "맑은고딕 14")  # 전체 폰트

    ontk = Label(count_item)  # 수량 레이블
    ontk.config(text="수량", width=10, relief="solid")
    ontk.pack(side="top", pady=10)

    amount = Entry(count_item)  # 수량 엔트리 go_enter 연결
    amount.config(width=10, relief="solid", borderwidth=0)
    amount.focus()
    amount.bind("<Return>", go_enter)
    amount.place(x=60, y=50)
    amount.pack()

    conf = Button(count_item, text="확인")  # 확인 버튼
    conf.config(width=10, height=3, command=go)  # go 연결
    # conf.place(x=30,y=200)
    conf.pack(side="bottom", pady=10)
    count_item.mainloop()
def l_refrech():
    del_t()
    i=0
    for rows in temp_sheet.iter_rows(min_row=2):
        if rows[2].value==None: #None으로 나오는걸 숫자 0으로 바꿔준다
            rows[2].value=0
        tree.insert('', 'end', text="", values=[rows[0].value,rows[1].value,rows[2].value,rows[3].value])
        print(rows[2].value)
        i+=1
def insert_tree2(sheet):
    del_t2()
    for row2 in sheet.iter_rows():
        if row2[0].value==None:
            print(type(row2[0].value))
            print("continue")
            continue
        elif (row2[4].value==None)|(row2[4].value=="None") | (row2[0].value!=" "):
            tree2.insert('', 'end', text="", values=[row2[0].value,row2[1].value,row2[2].value,row2[3].value," "])
            print("save with '' ")
        else:
            tree2.insert('', 'end', text="", values=[row2[0].value,row2[1].value,row2[2].value,row2[3].value,row2[4].value])
            print("else")
def c_click(event):
    def close():
        count_item.quit()
        count_item.destroy()
    def go():
        num = int(amount.get())  # 입력된 텍스트(수량)저장
        selectedItem = tree2.selection()[0]  # tree 선택한 위치 받기
        print(selectedItem)

        if (((tree2.item(selectedItem)['values'][2]) == None) | (tree2.item(selectedItem)['values'][2] < num)):
            messagebox.showinfo("", "수량보다 많이 입력하였습니다.")
        else:
            count=1
            temp_item=""
            for row3 in temp_sheet2.iter_rows():  # 중앙 수량 조절
                for cell in row3:
                    if cell.value == tree2.item(selectedItem)['values'][0]: #풀품명이 같으면
                        temp_item=row3[0].value #temp_item은 물품명
                        print(temp_item)
                        row3[2].value-=num #수량 조절
                        row3[3].value = row3[1].value * row3[2].value #가격 조정
                        if row3[2].value==0: #수량이 0이면
                            temp_sheet2.delete_rows(count,1) #삭제
                            info_file.save(info_xl)
                            print("sheet2.delete...")
                            break
                count+=1
            insert_tree2(temp_sheet2)
            print("lstart")
            for i in range (len(og_sheets)):
                for row2 in og_sheets[i]: #왼쪽 시트를 i를 통해서 특정
                    print(row2[0].value)
                    if row2[0].value==temp_item: #물품명이 같으면
                        print("equals")
                        print(row2[2].value)
                        if (row2[2].value==None) | (row2[2].value=="") | (row2[2].value=="0"): #수량이 없으면
                            row2[2].value=num #입력된 숫자만큼 추가
                        else:
                            row2[2].value+=num #수량 + 입력된 숫자만큼 추가
                        print(row2[2].value)
                        og_file.save(home) #왼쪽 시트 저장
                        l_refrech()
                        break
            close()
    def go_enter(event):
        num = int(amount.get())  # 입력된 텍스트(수량)저장
        selectedItem = tree2.selection()[0]  # tree 선택한 위치 받기
        print(selectedItem)

        if (((tree2.item(selectedItem)['values'][2]) == None) | (tree2.item(selectedItem)['values'][2] < num)):
            messagebox.showinfo("", "수량보다 많이 입력하였습니다.")
        else:
            count=1
            temp_item=""
            for row3 in temp_sheet2.iter_rows():  # 중앙 수량 조절
                for cell in row3:
                    if cell.value == tree2.item(selectedItem)['values'][0]:
                        temp_item=row3[0].value
                        print(temp_item)
                        row3[2].value-=num
                        row3[3].value = row3[1].value * row3[2].value
                        if row3[2].value==0:
                            temp_sheet2.delete_rows(count,1)
                            info_file.save(info_xl)
                            print("sheet2.delete...")
                            break
                count+=1
            insert_tree2(temp_sheet2)
            print("lstart")
            for i in range (len(og_sheets)):
                for row2 in og_sheets[i]:
                    print(row2[0].value)
                    if row2[0].value==temp_item:
                        print("equals")
                        print(row2[2].value)
                        if (row2[2].value==None) | (row2[2].value=="") |(row2[2].value=="0"):
                            row2[2].value=num
                        else:
                            row2[2].value+=num
                        print(row2[2].value)
                        og_file.save(home)
                        l_refrech()
                        break
            close()

    global temp_sheet
    print(temp_sheet)
    print(temp_sheet2)
    count_item = Tk()  # 불러오기 하면 나오는 화면

    count_item.geometry("200x150+500+300")  # 창의 크기
    count_item.title("수량 입력")  # 창의 제목
    count_item.option_add("*Font", "맑은고딕 14")  # 전체 폰트

    ontk = Label(count_item)  # 수량 레이블
    ontk.config(text="수량", width=10, relief="solid")
    ontk.pack(side="top", pady=10)

    amount = Entry(count_item)  # 수량 엔트리 go_enter 연결
    amount.config(width=10, relief="solid", borderwidth=0)
    amount.focus()
    amount.bind("<Return>", go_enter)
    amount.place(x=60, y=50)
    amount.pack()

    conf = Button(count_item, text="확인")  # 확인 버튼
    conf.config(width=10, height=3, command=go)  # go 연결
    # conf.place(x=30,y=200)
    conf.pack(side="bottom", pady=10)
    count_item.mainloop()
def set():
    print("set()")
    stop=False
    count=0
    #세트시트의 물품명과 다른 시트들의 물품명을 비교 --(리스트로 물품명 수량 저장)
    #같은 이름이 있으면 수량을 확인 --(리스트속 물품명으로 찾은 후 리스트속 수량으로 비교)
    #수량에 문제가 없으면 시트의 수량을 줄임 --(시트의 수량 - 리스트속 수량)

    #중앙 시트에 추가 전에 수량 확인 --(리스트속 물품명으로 찾은 후, or 같은게 없을 경우 추가)
    #수량 확인후 값 수정 -- (리스트속 값으로 수량 수정, 총 금액 수정)
    #세트속 물품 비교가 전부 끝날때 까지 반복

    #왼쪽 시트 저장
    #중앙 시트 저장
    #양쪽 리프레시

    for rows in set_sheet.iter_rows(): #세트에 row 길이만큼 반복
        if ((rows[0].value==None) | (rows[0].value==" ") | (rows[0].value=="")) & (stop==True):
            print("for rows in pinfo...if")
            break
        else:
            print("for rows in pinfo...else")
            for i in range(len(og_sheets)):  #
                for row2 in og_sheets[i]:  # 왼쪽 시트를 i를 통해서 특정
                    print(rows[0].value)
                    print(row2[0].value)
                    if (row2[0].value == rows[0].value) & (rows[0].value!="물품명"):  # 물품명이 같으면
                        if row2[2].value<rows[2].value: #세트의 수량보다 적으면
                            messagebox.showinfo("",(rows[0].value+"의 수량이 부족합니다."))
                            stop=True

                        else:
                            row2[2].value-=rows[2].value #왼쪽 수량 - 사용수량
                            print("touch left sheets")


    for rows in set_sheet.iter_rows():
        visit = False
        print("for rows in pinfo....2")
        print("rows",rows[0].value)
        print(stop)
        if stop==False: #수량에 문제가 없을경우
            for row3 in temp_sheet2.iter_rows():  # 중앙 수량 조절
                print("수량조절 시작",row3[0].value)
                if row3[0].value==rows[0].value: #물품명이 같으면
                    print("물품명이 같아서 값을 바꾸는 중",row3[0].value==rows[0].value)
                    row3[2].value+=rows[2].value
                    row3[3].value = row3[1].value * row3[2].value
                    visit=True
                if (visit==False) & (rows[0].value!="물품명"):
                    print("값이 달라서 어펜드 할 예정")
                    if (rows[3].value==None) | (rows[3].value==" ") | (rows[3].value==""):
                        print("단위가 비어있음")
                        temp_sheet2.append([rows[0].value, rows[1].value, rows[2].value,
                                            (rows[1].value * rows[2].value)," "])
                    else:
                        print("단위가 비어있지 않음")
                        print("else",rows[1].value, rows[2].value)
                        temp_sheet2.append([rows[0].value, rows[1].value, rows[2].value,(rows[1].value*rows[2].value),rows[3].value])

    if stop==False:
        print("stop==False")
        info_file.save(info_xl)
        og_file.save(home)  # 왼쪽 시트 저장
        l_refrech()
        insert_tree2(temp_sheet2)


if __name__ == "__main__":
    home = 'xl/전체물품리스트_세트저장용.xlsx'
    info_xl='xl/개인정보_물품.xlsx'

    og_file= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
    info_file=openpyxl.load_workbook(info_xl,data_only=True) #개인정보, 빈소별 물품정보 저장 공간(값으)

    info_sheets=[info_file['빈소1'],info_file['빈소2'],info_file['빈소3'],info_file['빈소5'],info_file['빈소6'],info_file['특101'],info_file['특102'],info_file['특201'],info_file['특202']]
    og_sheets=[og_file['식당판매'], og_file['매점판매'], og_file['장의용품'], og_file['상복'], og_file['기타']]
    set_sheet=og_file['세트']
    pinfo_sheet=info_file['개인정보'] #개인정보 출력용

    global temp_sheet
    global temp_sheet2
    global room

    win = tk.Tk() # 창 생성
    win.geometry("1200x720") # 창의 크기
    win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
    win.option_add("*Font", "맑은고딕 12") # 전체 폰트

    #-------------------------------------------------

    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three"],
                                displaycolumns=["one", "two", "three"], height=24)  # 3개 창 생성

    tree.column("#0", width=10, anchor="center")  # 1
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=90, anchor="center")  # 2
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")  # 3
    tree.heading("#2", text="단가", anchor="center")

    tree.column("#3", width=100, anchor="center")  # 4
    tree.heading("#3", text="수량", anchor="center")

    #-------------------------------------------------

    tree2 = tkinter.ttk.Treeview(win, columns=["one", "two", "three","four","five"],
                                displaycolumns=["one", "two", "three","four","five"], height=24)  # 4개 창 생성

    tree2.column("#0", width=10, anchor="center")  # 0
    tree2.heading("#0", text="", anchor="center")

    tree2.column("#1", width=90, anchor="center")  # 1
    tree2.heading("#1", text="물품명", anchor="center")

    tree2.column("#2", width=100, anchor="center")  # 2
    tree2.heading("#2", text="단가", anchor="center")

    tree2.column("#3", width=100, anchor="center")  # 3
    tree2.heading("#3", text="수량", anchor="center")

    tree2.column("#4", width=100, anchor="center")  # 4
    tree2.heading("#4", text="금액", anchor="center")

    tree2.column("#5", width=100, anchor="center")  # 5
    tree2.heading("#5", text="단위", anchor="center")

    #-------------------------------------------------

    시트1 = Button(win, text = "식당판매")
    시트1.config(width=7,height=2,command=btn1)
    시트1.place(x=50,y=150)

    시트2 = Button(win, text="장의용품")
    시트2.config(width=7, height=2, command=btn2)
    시트2.place(x=140, y=150)

    시트3 = Button(win, text = "장의용품")
    시트3.config(width=7,height=2,command=btn3)
    시트3.place(x=230,y=150)

    시트4 = Button(win, text = "상복")
    시트4.config(width=7,height=2,command=btn4)
    시트4.place(x=320,y=150)

    시트5 = Button(win, text = "기타")
    시트5.config(width=7,height=2,command=btn5)
    시트5.place(x=410,y=150)

    # 세트 = Button(win, text = "세트")
    # 세트.config(width=7,height=2, command=set)
    # 세트.place(x=500,y=150)

    # -------------------------------------------------

    불러오기 = Button(win, text = "불러오기")
    불러오기.config(width=7,height=2, command=recall)
    불러오기.place(x=600,y=10)

    # -------------------------------------------------

    tree.place(x=50,y=200)
    tree.bind("<Double-Button-1>",l_click)
    tree2.place(x=500,y=200)
    tree2.bind("<Double-Button-1>",c_click)

    # -------------------------------------------------

    상주성명 = tkinter.Label(win, text="상주성명", width=35, height=2, relief="solid")
    상주성명.place(x=50,y=60)

    빈소 = tkinter.Label(win, text="빈소", width=35, height=2, relief="solid")
    빈소.place(x=50, y=10)

    빈소기간 = tkinter.Label(win, text="빈소기간", width=35, height=2, relief="solid")
    빈소기간.place(x=50, y=110)

    win.mainloop() # 창 실행
