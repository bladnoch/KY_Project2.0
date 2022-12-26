import tkinter
from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk

home = 'xl/test.xlsx'
info_xl = 'xl/personal.xlsx'

og_file = openpyxl.load_workbook(home, data_only=True)  # 초기 시트 위치 저장(값으로)
info_file = openpyxl.load_workbook(info_xl, data_only=True)  # 개인정보, 빈소별 물품정보 저장 공간(값으)
trade_xl='xl/trade.xlsx'
trade_file=openpyxl.load_workbook(trade_xl,data_only=True)
info_sheets = [info_file['빈소1']]  # 지금은 하나만 사용하지만 빈소 창이 생기면 9개로 늘어날 것임
readsh=info_file['빈소1']  #Hong info_sheets에 리스트를 2개씩 사용해서 읽히지않아 추가 작성
readtrade=trade_file['거래명세서'] #거래명세서 시트 읽어오기

#def exsave():  # 빈소에 등록 된 물품명 읽어오기

row1 = []
for x in range(1, (readsh.max_row + 1)):  # 행의 끝까지 반복
    row1.append(readsh.cell(x, 1).value)  # row에 값 넣기 / 1을 바꾸면 열이 바뀜


    #print(*row)  # row 내의 목록 전체 출력 (테스트용)


#def exsave2():  # 거래명세서에 등록 된 물품명 읽어오기
row2 = []
for x in range(7, (readtrade.max_row + 1)):  # 7번행부터 (물품명 시작) 끝까지 반복
    row2.append(readtrade.cell(x, 3).value)  # C열부터 (물품명) row에 값넣기

'''def comtest(): #리스트 내의 값 하나하나 비교( 전체비교 x ) 반복문 이용 가능하면 이용
    for row1i, row2i in zip(row1, row2):
        print(row1i, row2i, row1i == row2i)

comtest()'''

def comte2(): #def comtest와 같은 동작 ( 전체비교 x ) 정상 작동 확인
    diff = [row1i == row2[i] for i, row1i in enumerate(row1)]
    print(diff)
    for i, row1i in enumerate(row1):
        print(row1i, row2[i], row1i == row2[i])

#comte2()

rowt = [value for value in row1 if value in row2] #리스트 두개 중 같은 값만 출력
print(rowt)

    #print(*row)  # row 내의 목록 전체 출력 (테스트용)

"""def compare():
    a = {exsave() == exsave2()}
    if a == "치즈김밥":
        print('일치')
    else:
        print('미일치')"""

#def ttt():

#C = set(row1) & set(row2) # set & 비교
#D = [i for i, j in zip(row1, row2)if i == j] #리스트 값 비교

#D = list(set(row1).intersection(row2)) # row1, row2의 중복 값 출력

#D = list(set(row1) - set(row2)) # row1, row2 서로 다른 값 출력
#E = list(set(row2) - set(row1)) # row1, row2 서로 다른 값 출력
#print(D,E)
    #if C is not None: #None 값 제외 후 받기
        #for x in C:
            #print(x)

#print(list(C))

#print(row1 == row2)





'''try:
    ttt()
except Exception as e:
    print(e)'''

'''def write():
    i = 7
    while i <= 20:
        if i == (readtrade.max_row + 1):
            break
    #while i >= readtrade.max_row + 1:
    readtrade.cell(i, 6).value = '2'
    i += 1'''
    #for x in range ((readtrade.max_row + 1), 6):
     #   readtrade.cell(x).value = '2' """



#write()
trade_file.save('xl/trade.xlsx')

#ttt()