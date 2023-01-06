import openpyxl

def og_rows(): #왼쪽 시트별 길이 저장 =>og_row(5개 기준)
    count = 0
    for i in range(len(og_sheets)):

        for rows in og_sheets[i].iter_rows():  # ws시트 row 길이를 count에 저장
            count += 1
        og_row[i]=count
        count=0


def print_sheet():
    count=0
    for i in range(5):
        print(count)
        for row in og_sheets[i].iter_rows(min_row=2):
            count+=1
            for cell in row:
                if cell.value=="치킨":
                    row[2].value=100
                    og_file.save(home)



home = '/Users/doungukkim/Desktop/workspace/gyproject/test/test.xlsx'
info_xl='/Users/doungukkim/Desktop/workspace/gyproject/test/개인정보_물품.xlsx'

og_file= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
info_file=openpyxl.load_workbook(info_xl,data_only=True) #개인정보, 빈소별 물품정보 저장 공간(값으)

inf_sheets=[info_file['빈소1']]
og_sheets=[og_file['식당판매'], og_file['매점판매'], og_file['장의용품'], og_file['상복'], og_file['기타']]  #시트 리스트에 저장 시트 이름 바꾸면 같이 바꿔야 함
og_row=['','','','',''] #길이 저장
og_l=[[],[],[],[],[]] #column 2개에 있는 cell info each list에 저장
new_l=[] #불러오거나 저장핳때 사용할 예정
temp_sheet=None

global og_p #왼쪽 목록 폼 출력용
global new_p #중앙 목록 폼 출력용
new_p=[]
global count

og_rows()
print_sheet()

# while (True):